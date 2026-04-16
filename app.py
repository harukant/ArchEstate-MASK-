# app.py - Aplicación principal de Flask para ArchEstate
# Esta aplicación maneja leads de propiedades, usuarios, profesionales y panel de administración.

# --- IMPORTS ---
import csv
import io
import os
import re
import sqlite3

from datetime import datetime
from functools import wraps
from io import StringIO

import openpyxl
import pytz

from flask import Flask, render_template, jsonify, request, session, redirect, url_for, Response, send_file, flash
from fpdf import FPDF
from werkzeug.security import generate_password_hash, check_password_hash

# --- CONFIGURACIÓN Y SETUP ---
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev_secret_key_123')  # Clave secreta para sesiones
DATABASE = os.path.join(os.path.dirname(__file__), 'database.db')


# --- FUNCIONES DE BASE DE DATOS ---
def get_db_connection():
    """Establece una conexión a la base de datos SQLite y configura el row factory para acceder por nombre de columna."""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Inicializa la base de datos de usuarios, leads, profesionales y auditoría si no existe"""
    with app.app_context():
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Tabla de Usuarios
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'client'
            )
        ''')
        
        # Tabla de Leads
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS leads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                type TEXT NOT NULL,
                property_type TEXT NOT NULL DEFAULT 'departamento',
                zone TEXT NOT NULL,
                budget TEXT NOT NULL,
                currency TEXT NOT NULL DEFAULT 'ARG',
                phone TEXT NOT NULL,
                email TEXT NOT NULL,
                floor_block TEXT DEFAULT '',
                usable_m2 INTEGER DEFAULT 0,
                elevator TEXT DEFAULT '',
                land_area INTEGER DEFAULT 0,
                built_area INTEGER DEFAULT 0,
                pool TEXT DEFAULT '',
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Actualizar esquemas antiguos sin las nuevas columnas
        cursor.execute('PRAGMA table_info(leads)')
        existing_columns = [row[1] for row in cursor.fetchall()]
        schema_updates = [
            ('property_type', "TEXT NOT NULL DEFAULT 'departamento'"),
            ('floor_block', "TEXT DEFAULT ''"),
            ('usable_m2', "INTEGER DEFAULT 0"),
            ('elevator', "TEXT DEFAULT ''"),
            ('land_area', "INTEGER DEFAULT 0"),
            ('built_area', "INTEGER DEFAULT 0"),
            ('pool', "TEXT DEFAULT ''")
        ]
        for column, column_type in schema_updates:
            if column not in existing_columns:
                cursor.execute(f"ALTER TABLE leads ADD COLUMN {column} {column_type}")

        # Tabla de Profesionales
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS professionals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                license TEXT NOT NULL UNIQUE,
                specialty TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending'
            )
        ''')

        # Tabla de Auditoría
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                action TEXT NOT NULL,
                target TEXT NOT NULL,
                admin TEXT NOT NULL
            )
        ''')
        
        # Insertar datos de prueba si las tablas están vacías
        cursor.execute('SELECT COUNT(*) FROM leads')
        if cursor.fetchone()[0] == 0:
            sample_leads = [
                ("Construcción Villa", "casa", "Marbella, Málaga", "1.2M - 1.5M", "EUR", "+34 612 345 678", "cliente1@ejemplo.com", "", 0, "", 1400, 800, "sí"),
                ("Compra Penthouse", "departamento", "Barcelona, Eixample", "850k - 1M", "EUR", "+34 699 887 766", "cliente2@ejemplo.com", "5º A", 120, "sí", 0, 0, ""),
                ("Remodelación Mansión", "casa", "Madrid, La Moraleja", "500k+", "EUR", "+34 655 443 322", "cliente3@ejemplo.com", "", 0, "", 2300, 1100, "no")
            ]
            cursor.executemany('INSERT INTO leads (type, property_type, zone, budget, currency, phone, email, floor_block, usable_m2, elevator, land_area, built_area, pool) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)', sample_leads)

        cursor.execute('SELECT COUNT(*) FROM professionals')
        if cursor.fetchone()[0] == 0:
            sample_pros = [
                ("Arq. Carlos Méndez", "COAM-12948", "Arquitectura Residencial", "approved"),
                ("Inmobiliaria Prime S.L.", "API-4402", "Lujo & Off-market", "approved"),
                ("Estudio Loft Design", "COAM-5521", "Interiorismo", "pending")
            ]
            cursor.executemany('INSERT INTO professionals (name, license, specialty, status) VALUES (?, ?, ?, ?)', sample_pros)
          
        # CREAMOS EL ADMIN POR DEFECTO (Ahora con su rol)
        cursor.execute('SELECT COUNT(*) FROM users')
        if cursor.fetchone()[0] == 0:
            cursor.execute('INSERT INTO users (username, hash, role) VALUES (?, ?, ?)', 
                          ('admin', generate_password_hash('admin123'), 'admin'))
        conn.commit()
        conn.close()


# --- DECORADORES ---
def login_required(f):
    """
    Decorador para proteger rutas que requieren autenticación.
    Si el usuario no está en la sesión, redirige al login.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    """
    Decorador para rutas que requieren rol de administrador.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        
        # Verificar rol en la base de datos
        conn = get_db_connection()
        user = conn.execute('SELECT role FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        conn.close()
        
        if not user or user['role'] != 'admin':
            flash('Acceso restringido: solo administradores pueden ingresar al panel de administración.')
            return redirect(url_for('index'))
        
        return f(*args, **kwargs)
    return decorated_function


def professional_required(f):
    """
    Decorador para rutas que requieren rol de profesional.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        
        # Verificar rol en la base de datos
        conn = get_db_connection()
        user = conn.execute('SELECT role FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        conn.close()
        
        if not user or user['role'] != 'professional':
            flash('Acceso denegado. Esta sección es solo para profesionales.')
            return redirect(url_for('index'))
        
        return f(*args, **kwargs)
    return decorated_function


# --- LÓGICA DE NEGOCIO (PYTHON) ---


def is_valid_email(email):
    """Lógica de validación de email en el servidor (más segura que JS)"""
    pattern = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
    return re.match(pattern, email) is not None


def log_action(action, target):
    """Registra una acción en la tabla de auditoría de la base de datos"""
    try:
        conn = get_db_connection()
        conn.execute('INSERT INTO audit_log (action, target, admin) VALUES (?, ?, ?)',
                     (action, target, session.get('username', 'sistema')))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error al registrar auditoría: {e}")


def convert_to_argentina_time(timestamp_str):
    """Convierte un timestamp UTC a hora de Argentina (UTC-3)"""
    if not timestamp_str:
        return timestamp_str
    try:
        # Parsear el timestamp (asumiendo que está en UTC)
        utc_time = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
        if utc_time.tzinfo is None:
            utc_time = pytz.UTC.localize(utc_time)
        
        # Convertir a Argentina (UTC-3)
        argentina_tz = pytz.timezone('America/Argentina/Buenos_Aires')
        argentina_time = utc_time.astimezone(argentina_tz)
        
        # Retornar en formato legible
        return argentina_time.strftime('%d/%m/%Y %H:%M:%S')
    except Exception as e:
        print(f"Error al convertir timestamp: {e}")
        return timestamp_str


def get_budget_stats_from_db():
    """Retorna estadísticas de presupuesto desde la base de datos"""
    conn = get_db_connection()
    
    # Total de leads
    total_leads = conn.execute('SELECT COUNT(*) FROM leads').fetchone()[0]
    
    # Leads por rango de presupuesto
    leads_by_budget = conn.execute(
        'SELECT budget, COUNT(*) as count FROM leads GROUP BY budget ORDER BY count DESC'
    ).fetchall()
    
    # Leads por moneda
    leads_by_currency = conn.execute(
        'SELECT currency, COUNT(*) as count FROM leads GROUP BY currency'
    ).fetchall()
    
    conn.close()
    
    return {
        'total_leads': total_leads,
        'by_budget': [{'label': r['budget'], 'value': r['count']} for r in leads_by_budget],
        'by_currency': [{'label': r['currency'], 'value': r['count']} for r in leads_by_currency],
    }

# --- RUTAS DE NAVEGACIÓN (VISTAS) ---

@app.route('/')
def index():
    return render_template('landing.html')


@app.route('/usuario')
@login_required
def user_view():
    # Solo rechazar a profesionales, admins y clientes pueden entrar
    conn = get_db_connection()
    user = conn.execute('SELECT role FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    conn.close()
    
    if user and user['role'] == 'professional':
        flash('Acceso denegado. Los profesionales no pueden acceder a esta sección.')
        return redirect(url_for('index'))
    
    return render_template('user.html')


@app.route('/profesional')
@professional_required
def professional_view():
    """Muestra los leads disponibles desde la base de datos"""
    conn = get_db_connection()
    
    # Obtener el usuario actual
    user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not user:
        conn.close()
        return redirect(url_for('index'))
    
    # Verificar si el profesional está aprobado
    professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
    if not professional or professional['status'] != 'approved':
        conn.close()
        return render_template('professional.html', leads=[], pending=True)
    
    # Si está aprobado, mostrar leads
    leads = conn.execute('SELECT * FROM leads ORDER BY timestamp, id DESC').fetchall()
    conn.close()
    
    # Convertir timestamps al huso horario de Argentina y a lista de diccionarios para Jinja2
    leads_list = []
    for lead in leads:
        lead_dict = dict(lead)
        lead_dict['timestamp'] = convert_to_argentina_time(lead_dict['timestamp'])
        leads_list.append(lead_dict)
    return render_template('professional.html', leads=leads_list, pending=False)


@app.route('/profesional/lead/<int:lead_id>')
@professional_required
def lead_detail(lead_id):
    conn = get_db_connection()
    user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not user:
        conn.close()
        return redirect(url_for('index'))

    professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
    if not professional or professional['status'] != 'approved':
        conn.close()
        return render_template('professional.html', leads=[], pending=True)

    lead = conn.execute('SELECT * FROM leads WHERE id = ?', (lead_id,)).fetchone()
    conn.close()

    if not lead:
        return redirect(url_for('professional_view'))

    lead_dict = dict(lead)
    lead_dict['timestamp'] = convert_to_argentina_time(lead_dict['timestamp'])
    return render_template('lead_detail.html', lead=lead_dict)


@app.route('/admin')
@admin_required
def admin_view():
    """Muestra profesionales y logs desde la base de datos"""
    conn = get_db_connection()
    professionals = conn.execute('SELECT * FROM professionals').fetchall()
    audit_logs = conn.execute('SELECT * FROM audit_log ORDER BY timestamp DESC').fetchall()
    conn.close()
    
    # Convertir timestamps de audit_log a UTC-3 (Argentina)
    audit_log_converted = []
    for log in audit_logs:
        log_dict = dict(log)
        log_dict['timestamp'] = convert_to_argentina_time(log_dict['timestamp'])
        audit_log_converted.append(log_dict)
    
    return render_template('admin.html', 
                           professionals=[dict(p) for p in professionals], 
                           audit_log=audit_log_converted)


# --- RUTAS DE API (LÓGICA DE DATOS) ---


# --- RUTA DE REGISTRO ---
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        raw_role = request.form.get('role', 'client')
        license_number = request.form.get('license', '').strip()

        # ✅ VALIDACIÓN DE CAMPOS OBLIGATORIOS
        if not username:
            flash('El nombre de usuario es requerido.')
            return redirect(url_for('register'))
        
        if not password or len(password) < 6:
            flash('La contraseña debe tener al menos 6 caracteres.')
            return redirect(url_for('register'))
        
        # 🛡️ VALIDACIÓN DE SEGURIDAD CRÍTICA (Backend)
        # Detectar intentos de inyectar 'admin' o roles no autorizados
        if raw_role == 'admin':
            print(f"⚠️ ALERTA DE SEGURIDAD: Intento de registro ilegal como admin por {username}")
            flash('Acceso denegado. Solo administradores pueden asignarse ese rol.')
            return redirect(url_for('register'))
        
        # Solo permitimos roles explícitamente definidos
        if raw_role in ['client', 'professional']:
            role = raw_role
        else:
            role = 'client'
        
        # ✅ VALIDACIÓN: Profesional requiere matrícula
        if role == 'professional' and not license_number:
            flash('El número de matrícula es requerido para profesionales.')
            return redirect(url_for('register'))

        conn = get_db_connection()
        try:
            # 1. Crear usuario (ahora usando el 'role' validado y seguro)
            cursor = conn.execute('INSERT INTO users (username, hash, role) VALUES (?, ?, ?)', 
                                 (username, generate_password_hash(password), role))
            
            # 2. Si es profesional, usar su matrícula real
            if role == 'professional':
                conn.execute('INSERT INTO professionals (name, license, specialty, status) VALUES (?, ?, ?, ?)',
                             (username, license_number, 'General', 'pending'))
            
            conn.commit()
            flash('Registro exitoso. Por favor, inicia sesión.')
            return redirect(url_for('login'))

        except sqlite3.IntegrityError as e:
            flash('El nombre de usuario ya está en uso. Por favor, elige otro.')
            return redirect(url_for('register'))
        except Exception as e:
            print(f"Error al registrar usuario: {e}")
            flash('Error al registrar. Por favor, intenta de nuevo.')
            return redirect(url_for('register'))
        finally:
            conn.close()

    return render_template('register.html')


# --- RUTA DE LOGIN ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        conn = get_db_connection()
        user = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        conn.close()

        # Verificar credenciales
        if user and check_password_hash(user['hash'], password):
            session.clear()
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']  # Guardar el rol en la sesión
            
            # Redirigir según el rol del usuario
            if user['role'] == 'admin':
                return redirect(url_for('admin_view'))
            elif user['role'] == 'professional':
                return redirect(url_for('professional_view'))
            else:  # client u otros roles
                return redirect(url_for('user_view'))
        
        # Si falla el login
        flash('Credenciales inválidas. Intente de nuevo.')
        return redirect(url_for('login'))

    return render_template('login.html')

# --- RUTA DE LOGOUT ---
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))


@app.route('/api/submit', methods=['POST'])
def submit_lead():
    """
    Envía una solicitud de propiedad.
    Solo usuarios autenticados pueden usar este endpoint.
    """
    data = request.json
    user_id = session.get('user_id')
    
    # --- VALIDACIÓN DE AUTENTICACIÓN ---
    if not user_id:
        return jsonify({
            "status": "error", 
            "message": "Debes estar registrado para enviar solicitudes."
        }), 401
    
    # Obtener datos del usuario desde la BD (no desde el formulario)
    conn = get_db_connection()
    user = conn.execute('SELECT id, username FROM users WHERE id = ?', (user_id,)).fetchone()
    
    if not user:
        conn.close()
        return jsonify({"status": "error", "message": "Sesión no válida"}), 401
    
    # --- GUARDADO EN BASE DE DATOS ---
    try:
        conn.execute('''
            INSERT INTO leads (
                type, property_type, zone, budget, currency, 
                phone, email, floor_block, usable_m2, elevator, 
                land_area, built_area, pool
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('type'),
            data.get('property_type', 'departamento'),
            data.get('zone'),
            data.get('budget'),
            data.get('currency', 'ARG'),
            data.get('phone'),
            data.get('email'),
            data.get('floor_block', ''),
            data.get('usable_m2', 0),
            data.get('elevator', ''),
            data.get('land_area', 0),
            data.get('built_area', 0),
            data.get('pool', '')
        ))
        conn.commit()
        conn.close()
        
        return jsonify({
            "status": "success", 
            "message": "Solicitud enviada con éxito. Los profesionales se contactarán contigo."
        })
    
    except Exception as e:
        conn.close()
        print(f"Error en BD: {e}")
        return jsonify({"status": "error", "message": "Error al procesar la solicitud."}), 500



@app.route('/api/budget-stats')
def budget_stats():
    """Retorna estadísticas de presupuesto en formato JSON"""
    stats = get_budget_stats_from_db()
    return jsonify(stats)


@app.route('/api/leads/export')
@professional_required
def export_leads_csv():
    """Genera y descarga un archivo CSV con todos los leads"""
    conn = get_db_connection()
    
    # Verificar si el profesional está aprobado
    user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not user:
        conn.close()
        return "Acceso denegado", 403
    
    professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
    if not professional or professional['status'] != 'approved':
        conn.close()
        return "Cuenta pendiente de aprobación", 403
    
    leads = conn.execute('SELECT id, type, zone, budget, currency, timestamp FROM leads ORDER BY timestamp DESC').fetchall()
    conn.close()

    def generate():
        data = StringIO()
        writer = csv.writer(data)
        
        # Escribir cabecera
        writer.writerow(['ID', 'Tipo Operacion', 'Zona', 'Presupuesto', 'Moneda', 'Fecha Registro (Argentina)'])
        yield data.getvalue()
        data.seek(0)
        data.truncate(0)

        # Escribir filas (convertir timestamps a UTC-3 Argentina)
        for lead in leads:
            timestamp_argentina = convert_to_argentina_time(lead['timestamp'])
            writer.writerow([lead['id'], lead['type'], lead['zone'], lead['budget'], lead['currency'], timestamp_argentina])
            yield data.getvalue()
            data.seek(0)
            data.truncate(0)

@app.route('/api/leads/export/xlsx')
@professional_required
def export_leads_xlsx():
    """Genera y descarga un archivo XLSX con todos los leads"""
    conn = get_db_connection()
    
    # Verificar si el profesional está aprobado
    user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not user:
        conn.close()
        return "Acceso denegado", 403
    
    professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
    if not professional or professional['status'] != 'approved':
        conn.close()
        return "Cuenta pendiente de aprobación", 403
    
    leads = conn.execute('SELECT id, type, zone, budget, currency, timestamp FROM leads ORDER BY timestamp DESC').fetchall()
    conn.close()

    # Crear workbook y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Cabeceras
    headers = ['ID', 'Tipo Operacion', 'Zona', 'Presupuesto', 'Moneda', 'Fecha Registro (Argentina)']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Datos
    for row_num, lead in enumerate(leads, 2):
        timestamp_argentina = convert_to_argentina_time(lead['timestamp'])
        ws.cell(row=row_num, column=1, value=lead['id'])
        ws.cell(row=row_num, column=2, value=lead['type'])
        ws.cell(row=row_num, column=3, value=lead['zone'])
        ws.cell(row=row_num, column=4, value=lead['budget'])
        ws.cell(row=row_num, column=5, value=lead['currency'])
        ws.cell(row=row_num, column=6, value=timestamp_argentina)

    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"leads_archestate_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/api/lead/<int:lead_id>/phone')
@professional_required
def get_lead_phone(lead_id):
    """
    Entrega el teléfono de un lead específico y audita la consulta.
    """
    conn = get_db_connection()
    
    # Verificar si el profesional está aprobado
    user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not user:
        conn.close()
        return jsonify({"error": "Acceso denegado"}), 403
    
    professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
    if not professional or professional['status'] != 'approved':
        conn.close()
        return jsonify({"error": "Cuenta pendiente de aprobación"}), 403
    
    lead = conn.execute('SELECT phone, type FROM leads WHERE id = ?', (lead_id,)).fetchone()
    
    if lead:
        # Auditar la consulta
        log_action("Consulta Teléfono", f"Lead ID: {lead_id} ({lead['type']})")
        conn.close()
        return jsonify({"phone": lead['phone']})
    
    conn.close()
    return jsonify({"error": "Lead no encontrado"}), 404


@app.route('/api/lead/<int:lead_id>/download')
@professional_required
def download_lead_pdf(lead_id):
    """Genera un PDF con los detalles del lead para descarga."""
    conn = get_db_connection()
    
    # Verificar si el profesional está aprobado
    user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not user:
        conn.close()
        return "Acceso denegado", 403
    
    professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
    if not professional or professional['status'] != 'approved':
        conn.close()
        return "Cuenta pendiente de aprobación", 403
    
    lead = conn.execute('SELECT * FROM leads WHERE id = ?', (lead_id,)).fetchone()
    conn.close()

    if not lead:
        return jsonify({"error": "Lead no encontrado"}), 404

    def safe_text(value):
        if value is None:
            return ''
        return str(value).replace('€', 'EUR').replace('—', '-').replace('–', '-')

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Título
    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 10, 'Detalle de Lead ArchEstate', ln=True, align='C')
    pdf.ln(6)

    # Contenido (Cambiamos multi_cell por cell)
    pdf.set_font('Helvetica', '', 12)
    
    # El parámetro ln=True hace que baje a la siguiente línea automáticamente
    pdf.cell(0, 8, f"ID Pedido: {safe_text(lead['id'])}", ln=True)
    pdf.cell(0, 8, f"Tipo de Operación: {safe_text(lead['type'])}", ln=True)
    pdf.cell(0, 8, f"Tipo de Propiedad: {safe_text(lead['property_type']).capitalize()}", ln=True)
    pdf.cell(0, 8, f"Zona: {safe_text(lead['zone'])}", ln=True)
    pdf.cell(0, 8, f"Presupuesto: {safe_text(lead['budget'])} ({safe_text(lead['currency'])})", ln=True)
    pdf.cell(0, 8, f"Teléfono: {safe_text(lead['phone'])}", ln=True)
    pdf.cell(0, 8, f"Email: {safe_text(lead['email'])}", ln=True)

    if safe_text(lead['property_type']).lower() == 'departamento':
        pdf.cell(0, 8, f"Piso / Bloque: {safe_text(lead['floor_block'])}", ln=True)
        pdf.cell(0, 8, f"Metros útiles: {safe_text(lead['usable_m2'])}", ln=True)
        pdf.cell(0, 8, f"Ascensor: {safe_text(lead['elevator'])}", ln=True)
    else:
        pdf.cell(0, 8, f"Superficie de terreno: {safe_text(lead['land_area'])} m²", ln=True)
        pdf.cell(0, 8, f"Superficie construida: {safe_text(lead['built_area'])} m²", ln=True)
        pdf.cell(0, 8, f"Piscina: {safe_text(lead['pool'])}", ln=True)

    pdf.cell(0, 8, f"Fecha registro: {safe_text(convert_to_argentina_time(lead['timestamp']))}", ln=True)

    # Generar el PDF
    pdf_output = pdf.output(dest='S')
    
    # 1. Asegurarnos de que sea sí o sí formato 'bytes' puro
    if isinstance(pdf_output, str):
        pdf_bytes = pdf_output.encode('latin-1')
    else:
        pdf_bytes = bytes(pdf_output) # Convertir bytearray a bytes puros

    # 2. Crear un archivo virtual en la memoria (Buffer)
    buffer = io.BytesIO(pdf_bytes)
    buffer.seek(0) # Volver el puntero al inicio del archivo virtual
    
    filename = f"lead_{lead['id']}.pdf"

    # 3. Usar send_file de Flask, que calcula el Content-Length automáticamente y sin errores
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename, # Si usan una versión vieja de Flask, cambien download_name por attachment_filename
        mimetype='application/pdf'
    )


@app.route('/api/admin/professional/<int:pro_id>/status', methods=['POST'])
@admin_required
def update_pro_status(pro_id):
    """Actualiza el estado de un profesional en la BD y registra la acción"""
    data = request.json
    new_status = data.get('status')
    
    if new_status not in ['approved', 'rejected']:
        return jsonify({"error": "Estado no válido"}), 400
        
    conn = get_db_connection()
    pro = conn.execute('SELECT name FROM professionals WHERE id = ?', (pro_id,)).fetchone()
    
    if pro:
        conn.execute('UPDATE professionals SET status = ? WHERE id = ?', (new_status, pro_id))
        conn.commit()
        
        action = "Aprobación" if new_status == 'approved' else "Rechazo"
        log_action(action, pro['name'])
        conn.close()
        return jsonify({"status": "success", "message": f"Profesional {action.lower()} correctamente"})
        
    conn.close()
    return jsonify({"error": "Profesional no encontrado"}), 404


@app.route('/api/admin/stats')
@login_required
def admin_stats():
    """Retorna estadísticas agregadas para el dashboard del admin"""
    conn = get_db_connection()

    # Total de leads
    total_leads = conn.execute('SELECT COUNT(*) FROM leads').fetchone()[0]

    # Leads por tipo de operación
    leads_by_type = conn.execute(
        'SELECT type, COUNT(*) as count FROM leads GROUP BY type ORDER BY count DESC'
    ).fetchall()

    # Leads por zona (top 5)
    leads_by_zone = conn.execute(
        'SELECT zone, COUNT(*) as count FROM leads GROUP BY zone ORDER BY count DESC LIMIT 5'
    ).fetchall()

    # Leads por presupuesto
    leads_by_budget = conn.execute(
        'SELECT budget, COUNT(*) as count FROM leads GROUP BY budget ORDER BY count DESC'
    ).fetchall()

    # Leads por mes (últimos 6 meses)
    leads_by_month = conn.execute('''
        SELECT strftime('%Y-%m', timestamp) as month, COUNT(*) as count
        FROM leads
        GROUP BY month
        ORDER BY month DESC
        LIMIT 6
    ''').fetchall()

    # Estado de profesionales
    pros_stats = conn.execute(
        'SELECT status, COUNT(*) as count FROM professionals GROUP BY status'
    ).fetchall()

    # Total de usuarios por rol
    users_by_role = conn.execute(
        'SELECT role, COUNT(*) as count FROM users GROUP BY role'
    ).fetchall()

    # Acciones del log de auditoría
    audit_actions = conn.execute(
        'SELECT action, COUNT(*) as count FROM audit_log GROUP BY action ORDER BY count DESC'
    ).fetchall()

    conn.close()

    return jsonify({
        'total_leads': total_leads,
        'leads_by_type': [{'label': r['type'], 'value': r['count']} for r in leads_by_type],
        'leads_by_zone': [{'label': r['zone'], 'value': r['count']} for r in leads_by_zone],
        'leads_by_budget': [{'label': r['budget'], 'value': r['count']} for r in leads_by_budget],
        'leads_by_month': [{'label': r['month'], 'value': r['count']} for r in reversed(leads_by_month)],
        'pros_stats': [{'label': r['status'], 'value': r['count']} for r in pros_stats],
        'users_by_role': [{'label': r['role'], 'value': r['count']} for r in users_by_role],
        'audit_actions': [{'label': r['action'], 'value': r['count']} for r in audit_actions],
    })



# Inicializar la base de datos al arrancar
init_db()


if __name__ == '__main__':
    app.run(debug=True)



  # { "workspaceRoot": "file:///vsls:/", "fileUri": "file:///vsls:/app.py" }  



